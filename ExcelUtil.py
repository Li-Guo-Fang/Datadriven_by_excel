from openpyxl import load_workbook

class ParseExcel(object):

    def __init__(self, excelPath, sheetName):
        self.wb = load_workbook(excelPath)
        self.sheet = self.wb.get_sheet_by_name(sheetName)
        self.maxRowNum = self.sheet.max_row

    def getDatasFromSheet(self):
        # 用于存放从工作表中读取出来的数据
        dataList = []
        for line in self.sheet.rows:
            tmpList = []
            tmpList.append(line[1].value)     #第一个单元格的值
            tmpList.append(line[2].value)     #第二个单元格的值
            dataList.append(tmpList)
        return dataList[1:]

if __name__ == '__main__':
    excelPath = u'D:\data.xlsx'
    sheetName = u"搜索数据表"
    pe = ParseExcel(excelPath, sheetName)
    print (pe.getDatasFromSheet())
    for i in pe.getDatasFromSheet():
        print (i[0], i[1])